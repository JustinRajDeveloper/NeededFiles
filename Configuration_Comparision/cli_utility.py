def output_excel(comparison_results, security_issues, output_file=None, only_mismatches=False):
    """Output results to Excel with security analysis"""
    if not output_file:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"properties_comparison_{timestamp}.xlsx"
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Security Summary sheet
        create_security_summary_sheet(writer, security_issues)
        
        # Main Summary sheet
        summary_data = []
        for env_config, env_data in comparison_results.#!/usr/bin/env python3
"""
CLI utility for microservice properties comparison
Usage: python cli_compare.py --repo-url https://github.com/org/repo --microservices service1,service2,service3
"""

import argparse
import json
import sys
from app import GitHubAPIPropertyComparator
import pandas as pd
from datetime import datetime
import os

def main():
    parser = argparse.ArgumentParser(description='Compare microservice properties')
    parser.add_argument('--repo-url', required=True, help='GitHub repository URL')
    parser.add_argument('--microservices', required=True, 
                       help='Comma-separated list of microservice names')
    parser.add_argument('--github-token', help='GitHub personal access token')
    parser.add_argument('--branch', default='main', help='Git branch to scan (default: main)')
    parser.add_argument('--output', choices=['console', 'json', 'excel', 'csv'], 
                       default='console', help='Output format')
    parser.add_argument('--output-file', help='Output file path (for json, excel, csv)')
    parser.add_argument('--only-mismatches', action='store_true', 
                       help='Show only mismatched properties')
    parser.add_argument('--environment', help='Filter by specific environment (e.g., dev_dev1)')
    parser.add_argument('--verbose', '-v', action='store_true', help='Verbose output')
    
    args = parser.parse_args()
    
    # Parse microservices list
    microservices = [ms.strip() for ms in args.microservices.split(',')]
    
    if args.verbose:
        print(f"Repository: {args.repo_url}")
        print(f"Microservices: {microservices}")
        print(f"Branch: {args.branch}")
        print(f"Output format: {args.output}")
        print("-" * 50)
    
    try:
        # Initialize comparator
        comparator = GitHubAPIPropertyComparator(args.github_token)
        
        # Scan microservices
        if args.verbose:
            print("Scanning microservices...")
        
        all_data = comparator.scan_selected_microservices(
            args.repo_url, microservices, args.branch
        )
        
        if not all_data:
            print("ERROR: No data found for the specified microservices", file=sys.stderr)
            sys.exit(1)
        
        # Compare environments
        if args.verbose:
            print("Comparing environments...")
        
        comparison_results = comparator.compare_environments(all_data)
        
        # Analyze security issues
        if args.verbose:
            print("Analyzing security issues...")
        
        security_issues = comparator.analyze_security_issues(all_data)
        
        # Filter by environment if specified
        if args.environment:
            if args.environment in comparison_results:
                comparison_results = {args.environment: comparison_results[args.environment]}
            else:
                print(f"ERROR: Environment '{args.environment}' not found", file=sys.stderr)
                available = list(comparison_results.keys())
                print(f"Available environments: {available}", file=sys.stderr)
                sys.exit(1)
        
        # Generate output
        if args.output == 'console':
            output_console(comparison_results, security_issues, args.only_mismatches, args.verbose)
        elif args.output == 'json':
            output_json(comparison_results, security_issues, args.output_file, args.only_mismatches)
        elif args.output == 'excel':
            output_excel(comparison_results, security_issues, args.output_file, args.only_mismatches)
        elif args.output == 'csv':
            output_csv(comparison_results, security_issues, args.output_file, args.only_mismatches)
        
        if args.verbose:
            print_summary(comparison_results, security_issues)
            
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        if args.verbose:
            import traceback
            traceback.print_exc()
        sys.exit(1)

def output_console(comparison_results, security_issues, only_mismatches=False, verbose=False):
    """Output results to console"""
    # First, show security issues if any
    show_security_summary(security_issues)
    
    for env_config, env_data in comparison_results.items():
        print(f"\n{'='*60}")
        print(f"ENVIRONMENT: {env_config.upper()}")
        print(f"{'='*60}")
        print(f"Microservices: {', '.join(env_data['microservices'])}")
        print(f"Total Properties: {env_data['total_properties']}")
        print(f"Matched: {env_data['matched_count']}")
        print(f"Mismatched: {env_data['mismatched_count']}")
        
        # Show mismatched properties
        if env_data['mismatched_count'] > 0:
            print(f"\n❌ MISMATCHED PROPERTIES ({env_data['mismatched_count']}):")
            print("-" * 40)
            for prop_key, ms_values in env_data['mismatched'].items():
                print(f"\n🔑 {prop_key}")
                for ms, value in ms_values.items():
                    print(f"   {ms:20} = {value}")
        
        # Show matched properties if not filtering
        if not only_mismatches and env_data['matched_count'] > 0:
            print(f"\n✅ MATCHED PROPERTIES ({env_data['matched_count']}):")
            print("-" * 40)
            for prop_key, ms_values in env_data['matching'].items():
                value = list(ms_values.values())[0]  # All values are the same
                print(f"🔑 {prop_key} = {value}")

def show_security_summary(security_issues):
    """Show security issues summary"""
    total_secrets = sum(len(issues) for issues in security_issues['hardcoded_secrets'].values())
    total_insecure = sum(len(issues) for issues in security_issues['insecure_protocols'].values())
    total_weak = sum(len(issues) for issues in security_issues['weak_configurations'].values())
    
    if total_secrets + total_insecure + total_weak == 0:
        print("🔒 SECURITY: No security issues detected")
        return
    
    print(f"\n{'🚨 SECURITY ANALYSIS'}")
    print("=" * 60)
    
    if total_secrets > 0:
        print(f"🔴 CRITICAL: {total_secrets} hardcoded secrets found")
        for env, issues in security_issues['hardcoded_secrets'].items():
            for issue in issues:
                masked_value = mask_value(issue['value'])
                print(f"   {env} | {issue['microservice']} | {issue['property']} = {masked_value}")
    
    if total_insecure > 0:
        print(f"🟡 MEDIUM: {total_insecure} insecure protocol configurations")
        for env, issues in security_issues['insecure_protocols'].items():
            for issue in issues:
                print(f"   {env} | {issue['microservice']} | {issue['property']} = {issue['value']}")
    
    if total_weak > 0:
        print(f"🔵 LOW-MEDIUM: {total_weak} weak security configurations")
        for env, issues in security_issues['weak_configurations'].items():
            for issue in issues:
                print(f"   {env} | {issue['microservice']} | {issue['property']} = {issue['value']}")

def mask_value(value):
    """Mask sensitive values for console output"""
    if len(value) > 8:
        return value[:4] + '*' * (len(value) - 8) + value[-4:]
    elif len(value) > 4:
        return value[:2] + '*' * (len(value) - 4) + value[-2:]
    else:
        return '*' * len(value)

def output_json(comparison_results, output_file=None, only_mismatches=False):
    """Output results to JSON"""
    if only_mismatches:
        filtered_results = {}
        for env, data in comparison_results.items():
            if data['mismatched_count'] > 0:
                filtered_results[env] = {
                    'microservices': data['microservices'],
                    'mismatched': data['mismatched'],
                    'mismatched_count': data['mismatched_count']
                }
        output_data = filtered_results
    else:
        output_data = comparison_results
    
    json_output = json.dumps(output_data, indent=2)
    
    if output_file:
        with open(output_file, 'w') as f:
            f.write(json_output)
        print(f"JSON output saved to: {output_file}")
    else:
        print(json_output)

def output_excel(comparison_results, security_issues, output_file=None, only_mismatches=False):
    """Output results to Excel with security analysis"""
    if not output_file:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"properties_comparison_{timestamp}.xlsx"
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Security Summary sheet (first sheet for visibility)
        create_security_summary_sheet(writer, security_issues)
        
        # Main Summary sheet
        summary_data = []
        for env_config, env_data in comparison_results.items():
            # Count security issues for this environment
            env_security_count = 0
            for issue_type in security_issues.values():
                if env_config in issue_type:
                    env_security_count += len(issue_type[env_config])
            
            summary_data.append({
                'Environment': env_config,
                'Microservices': len(env_data['microservices']),
                'Total Properties': env_data['total_properties'],
                'Matched': env_data['matched_count'],
                'Mismatched': env_data['mismatched_count'],
                'Security Issues': env_security_count,
                'Match %': f"{(env_data['matched_count'] / max(env_data['total_properties'], 1)) * 100:.1f}%",
                'Security Status': '🔴 CRITICAL' if env_security_count > 0 else '🟢 SECURE'
            })
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Detailed sheets for each environment
        for env_config, env_data in comparison_results.items():
            sheet_data = []
            microservices = env_data['microservices']
            
            # Add mismatched properties
            for prop_key, ms_values in env_data['mismatched'].items():
                row = {'Property': prop_key, 'Status': 'MISMATCH', 'Security_Risk': ''}
                
                # Check if this property has security issues
                security_risk = check_property_security_risk(prop_key, ms_values, security_issues, env_config)
                if security_risk:
                    row['Security_Risk'] = security_risk
                
                for ms in microservices:
                    row[ms] = ms_values.get(ms, 'N/A')
                sheet_data.append(row)
            
            # Add matched properties (if not filtering)
            if not only_mismatches:
                for prop_key, ms_values in env_data['matching'].items():
                    row = {'Property': prop_key, 'Status': 'MATCH', 'Security_Risk': ''}
                    
                    # Check if this property has security issues
                    security_risk = check_property_security_risk(prop_key, ms_values, security_issues, env_config)
                    if security_risk:
                        row['Security_Risk'] = security_risk
                    
                    for ms in microservices:
                        row[ms] = ms_values.get(ms, 'N/A')
                    sheet_data.append(row)
            
            if sheet_data:
                df = pd.DataFrame(sheet_data)
                sheet_name = env_config.replace('/', '_')[:31]  # Excel sheet name limit
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Apply conditional formatting
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                apply_security_formatting(workbook, worksheet, df)
    
    print(f"Excel output with security analysis saved to: {output_file}")

def create_security_summary_sheet(writer, security_issues):
    """Create a dedicated security summary sheet"""
    security_data = []
    
    # Collect all security issues
    for issue_category, category_data in security_issues.items():
        category_name = issue_category.replace('_', ' ').title()
        for env, issues in category_data.items():
            for issue in issues:
                security_data.append({
                    'Environment': env,
                    'Microservice': issue['microservice'],
                    'Property': issue['property'],
                    'Issue_Category': category_name,
                    'Issue_Type': issue['issue_type'],
                    'Severity': issue['severity'],
                    'Value': mask_value_for_excel(issue['value']),
                    'Recommendation': get_security_recommendation(issue['issue_type'])
                })
    
    if security_data:
        security_df = pd.DataFrame(security_data)
        security_df.to_excel(writer, sheet_name='🔒 Security Issues', index=False)
        
        # Add security statistics
        stats_data = []
        total_critical = sum(1 for item in security_data if item['Severity'] == 'HIGH')
        total_medium = sum(1 for item in security_data if item['Severity'] == 'MEDIUM')
        total_low = sum(1 for item in security_data if item['Severity'] == 'LOW')
        
        stats_data.append(['🔴 CRITICAL Issues', total_critical])
        stats_data.append(['🟡 MEDIUM Issues', total_medium])
        stats_data.append(['🔵 LOW Issues', total_low])
        stats_data.append(['📊 Total Issues', len(security_data)])
        
        stats_df = pd.DataFrame(stats_data, columns=['Category', 'Count'])
        
        # Write to a separate area in the same sheet
        start_row = len(security_df) + 3
        stats_df.to_excel(writer, sheet_name='🔒 Security Issues', 
                         startrow=start_row, index=False)
    else:
        # Create empty sheet with good news
        no_issues_df = pd.DataFrame([['🟢 No security issues detected!']], 
                                   columns=['Security Status'])
        no_issues_df.to_excel(writer, sheet_name='🔒 Security Issues', index=False)

def check_property_security_risk(prop_key, ms_values, security_issues, env_config):
    """Check if a property has security risks"""
    for issue_category, category_data in security_issues.items():
        if env_config in category_data:
            for issue in category_data[env_config]:
                if issue['property'] == prop_key:
                    severity_emoji = {'HIGH': '🔴', 'MEDIUM': '🟡', 'LOW': '🔵'}
                    return f"{severity_emoji.get(issue['severity'], '⚪')} {issue['issue_type']}"
    return ''

def mask_value_for_excel(value):
    """Mask sensitive values for Excel output"""
    if 'password' in value.lower() or 'secret' in value.lower() or 'key' in value.lower():
        if len(value) > 8:
            return value[:4] + '*' * (len(value) - 8) + value[-4:]
        elif len(value) > 4:
            return value[:2] + '*' * (len(value) - 4) + value[-2:]
        else:
            return '*' * len(value)
    return value

def get_security_recommendation(issue_type):
    """Get security recommendation for each issue type"""
    recommendations = {
        'password': 'Use environment variables or secret management service',
        'secret': 'Externalize to HashiCorp Vault, AWS Secrets Manager, or similar',
        'key': 'Store in secure key management system',
        'token': 'Use OAuth2 or JWT with external token provider',
        'credential': 'Use service accounts or external identity providers',
        'http_urls': 'Replace HTTP with HTTPS for secure communication',
        'unencrypted_db': 'Enable SSL/TLS for database connections',
        'insecure_protocols': 'Use secure protocol versions (HTTPS, TLS 1.2+)',
        'ssl_disabled': 'Enable SSL/TLS encryption in production',
        'debug_enabled': 'Disable debug mode in production environments',
        'weak_encryption': 'Use strong encryption algorithms (AES-256, SHA-256+)',
        'permissive_cors': 'Restrict CORS to specific allowed origins',
        'no_authentication': 'Enable authentication mechanisms'
    }
    return recommendations.get(issue_type, 'Review and secure this configuration')

def apply_security_formatting(workbook, worksheet, df):
    """Apply conditional formatting to highlight security issues"""
    from openpyxl.styles import PatternFill, Font
    
    # Define colors for different security levels
    critical_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
    medium_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
    low_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    
    # Apply formatting based on security risk column
    security_col_idx = None
    for idx, col in enumerate(df.columns):
        if col == 'Security_Risk':
            security_col_idx = idx + 1  # Excel is 1-indexed
            break
    
    if security_col_idx:
        for row_idx, row in enumerate(df.itertuples(), start=2):  # Start from row 2 (after header)
            security_value = getattr(row, 'Security_Risk', '')
            if '🔴' in str(security_value):
                for col_idx in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = critical_fill
            elif '🟡' in str(security_value):
                for col_idx in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = medium_fill
            elif '🔵' in str(security_value):
                for col_idx in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = low_fill

def print_summary(comparison_results, security_issues):
    """Print summary statistics including security"""
    total_envs = len(comparison_results)
    total_mismatches = sum(env['mismatched_count'] for env in comparison_results.values())
    total_matches = sum(env['matched_count'] for env in comparison_results.values())
    total_properties = sum(env['total_properties'] for env in comparison_results.values())
    
    # Security statistics
    total_critical = sum(len(issues) for issues in security_issues['hardcoded_secrets'].values())
    total_medium = sum(len(issues) for issues in security_issues['insecure_protocols'].values())
    total_weak = sum(len(issues) for issues in security_issues['weak_configurations'].values())
    total_security_issues = total_critical + total_medium + total_weak
    
    print(f"\n{'='*60}")
    print("SUMMARY")
    print(f"{'='*60}")
    print(f"Environments processed: {total_envs}")
    print(f"Total properties: {total_properties}")
    print(f"Total matches: {total_matches}")
    print(f"Total mismatches: {total_mismatches}")
    
    if total_properties > 0:
        match_percentage = (total_matches / total_properties) * 100
        print(f"Match percentage: {match_percentage:.1f}%")
    
    print(f"\n🔒 SECURITY SUMMARY:")
    print(f"Total security issues: {total_security_issues}")
    print(f"🔴 Critical (hardcoded secrets): {total_critical}")
    print(f"🟡 Medium (insecure protocols): {total_medium}")
    print(f"🔵 Low-Medium (weak configs): {total_weak}")
    
    if total_security_issues > 0:
        print(f"\n⚠️  Found {total_security_issues} security issues that need immediate attention!")
        if total_critical > 0:
            print(f"🚨 URGENT: {total_critical} hardcoded secrets must be externalized!")
    else:
        print(f"\n✅ No security issues detected!")
    
    if total_mismatches > 0:
        print(f"\n⚠️  Found {total_mismatches} property mismatches that need attention!")
    else:
        print(f"\n✅ All properties match across microservices!")

def output_csv(comparison_results, output_file=None, only_mismatches=False):
    """Output results to CSV"""
    if not output_file:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"properties_comparison_{timestamp}.csv"
    
    all_data = []
    
    for env_config, env_data in comparison_results.items():
        microservices = env_data['microservices']
        
        # Add mismatched properties
        for prop_key, ms_values in env_data['mismatched'].items():
            for ms, value in ms_values.items():
                all_data.append({
                    'Environment': env_config,
                    'Property': prop_key,
                    'Microservice': ms,
                    'Value': value,
                    'Status': 'MISMATCH'
                })
        
        # Add matched properties (if not filtering)
        if not only_mismatches:
            for prop_key, ms_values in env_data['matching'].items():
                for ms, value in ms_values.items():
                    all_data.append({
                        'Environment': env_config,
                        'Property': prop_key,
                        'Microservice': ms,
                        'Value': value,
                        'Status': 'MATCH'
                    })
    
    df = pd.DataFrame(all_data)
    df.to_csv(output_file, index=False)
    print(f"CSV output saved to: {output_file}")

def print_summary(comparison_results):
    """Print summary statistics"""
    total_envs = len(comparison_results)
    total_mismatches = sum(env['mismatched_count'] for env in comparison_results.values())
    total_matches = sum(env['matched_count'] for env in comparison_results.values())
    total_properties = sum(env['total_properties'] for env in comparison_results.values())
    
    print(f"\n{'='*60}")
    print("SUMMARY")
    print(f"{'='*60}")
    print(f"Environments processed: {total_envs}")
    print(f"Total properties: {total_properties}")
    print(f"Total matches: {total_matches}")
    print(f"Total mismatches: {total_mismatches}")
    
    if total_properties > 0:
        match_percentage = (total_matches / total_properties) * 100
        print(f"Match percentage: {match_percentage:.1f}%")
    
    if total_mismatches > 0:
        print(f"\n⚠️  Found {total_mismatches} property mismatches that need attention!")
    else:
        print(f"\n✅ All properties match across microservices!")

if __name__ == '__main__':
    main()

# Example usage:
"""
# Basic comparison
python cli_compare.py --repo-url https://github.com/myorg/microservices \
    --microservices user-service,payment-service,notification-service

# Only show mismatches in JSON format
python cli_compare.py --repo-url https://github.com/myorg/microservices \
    --microservices user-service,payment-service \
    --output json --only-mismatches --output-file mismatches.json

# Export to Excel with GitHub token for private repo
python cli_compare.py --repo-url https://github.com/myorg/private-microservices \
    --microservices service1,service2,service3 \
    --github-token ghp_xxxxxxxxxxxx \
    --output excel --output-file comparison_report.xlsx

# Filter specific environment
python cli_compare.py --repo-url https://github.com/myorg/microservices \
    --microservices user-service,payment-service \
    --environment dev_dev2 --verbose

# CI/CD Integration example
python cli_compare.py --repo-url $REPO_URL \
    --microservices $MICROSERVICES_LIST \
    --github-token $GITHUB_TOKEN \
    --only-mismatches --output json | \
    jq '.[] | select(.mismatched_count > 0)' && exit 1 || exit 0
"""